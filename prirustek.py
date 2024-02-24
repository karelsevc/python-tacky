import pandas as pd
from openpyxl_image_loader import SheetImageLoader
from openpyxl import load_workbook
import csv
import os
from datetime import date
#    Leden_2023_tata_B2007.xlsx  mensi zjednodusena verze 8 tacku ve slozce
#    Leden_2023_ok2007.xlsx plna verze je ulozena v LibreCalc do verze Excel 2007
jmeno_souboru = 'prirustek.xlsx'

celkem_row = 0  # pocet radku pres vsechny slozky

today = date.today()
year = today.year
month = today.month
day = today.day

# vycisteni adresaru ObrKmen*
akt_adr = os.getcwd()
adresare = ['ObrKmenC', 'ObrKmenD', 'ObrKmenE']
for adr_name in adresare:
    adr_dst = akt_adr + '/' + adr_name
    soubory = os.listdir(adr_dst)
    for file_name in soubory:   # smazani souboru v adresari ObrKmen*
        os.remove(adr_dst + '\\' + file_name)

# Vypsání počtu řádků v tabulce
with open('prirustek.csv', 'w', newline='', encoding="utf-8") as csvfile:
    writer = csv.writer(csvfile, delimiter='|')
    cestaC = 'ObrKmenC/'
    cestaD = 'ObrKmenD/'
    cestaE = 'ObrKmenE/'
# iterace pres vsechny sheet
    dfs = pd.ExcelFile(jmeno_souboru)
    for sheet_name in ['A']:
        df = pd.read_excel(jmeno_souboru, header=1, sheet_name=sheet_name)
        print('--------------------------------------------------------------')
        poc_radku = len(df)
        celkem_row = celkem_row + poc_radku
        print(f"{sheet_name} Pocet radku v tabulce je : {poc_radku}")
        pxl_doc = load_workbook(jmeno_souboru)
        sheet = pxl_doc[sheet_name]
#       print(sheet.max_row)  Poznamka: nepocita spravne asi chyba v knihovne dale je ok
        image_loader = SheetImageLoader(sheet)
        radka = 2
# Iterace přes všechny řádky
        for index, row in df.iterrows():
            img_soubor = str(radka)
            obrazekC = f"{cestaC}Kmen_{sheet_name}_{img_soubor}.jpg"
            obrazekD = None
            obrazekE = None
            countObr = radka + 1   # ma jine cislovani je od 1
            img_soubor = str(countObr)
            radka = radka + 1   # inkrementuj radku v sesitu
            precti = 'C' + img_soubor   # prectu sloupec 'C1' 'C2' 'C3' atd. podle citace
#            print(precti)
            image = image_loader.get(precti)
            image.thumbnail((90, 90))
            image.save(obrazekC)
            # Sloupec D
            precti = 'D' + img_soubor  # prectu sloupec 'D1' 'D2' 'D3' atd. podle citace
            if image_loader.image_in(precti):
                try:
                    image = image_loader.get(precti)
                    obrazekD = f"{cestaD}Obrazky_D_{sheet_name}_{img_soubor}.jpg"
                    image.thumbnail((90, 90))
                    image.save(obrazekD)
#                    print('Buňka  obsahuje obrázek')
                except ValueError:
                    print('Buňka  neobsahuje obrázek')
            precti = 'E' + img_soubor  # prectu sloupec 'D1' 'D2' 'D3' atd. podle citace
            if image_loader.image_in(precti):
                try:
                    image = image_loader.get(precti)
                    obrazekE = f"{cestaE}Obrazky_E_{sheet_name}_{img_soubor}.jpg"
                    image.thumbnail((90, 90))
                    image.save(obrazekE)
#                    print('Buňka  obsahuje obrázek')
                except ValueError:
                    print('Buňka  neobsahuje obrázek')
            row_new = f"{row[0]}| {row[1]}| {obrazekC}| {obrazekD}| "\
                      f"{obrazekE}| {row[5]}| {row[6]}| {row[7]}| "\
                      f"{day}.{month}.{year}"
            row = row_new
#          print(row)   vytiskni zpracovanou radku
            rowtuple = (row.split("| "))
            print(rowtuple)
            writer.writerow(rowtuple)

print(f"Celkem radku: {celkem_row}")
